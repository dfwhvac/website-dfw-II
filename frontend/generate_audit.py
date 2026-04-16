import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DFW HVAC Site Audit"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="003153", end_color="003153", fill_type="solid")
warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
null_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
section_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
section_font = Font(bold=True, size=11)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Headers
headers = ["Webpage (URL)", "Meta Description", "Status", "Live URL"]
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 80
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 55

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border

# Data
pages = [
    # Section: Static Pages
    ("SECTION", "Static Pages", "", "", ""),
    ("/", "Family-owned HVAC contractor serving Dallas-Fort Worth. AC and heating diagnosis, repair, installation and maintenance. Call (972) 777-COOL.", "Published", "https://dfwhvac.com/"),
    ("/about", "Learn about DFW HVAC - a three-generation family commitment to trustworthy, high-quality HVAC service in Dallas-Fort Worth.", "Published", "https://dfwhvac.com/about"),
    ("/contact", "Contact DFW HVAC for expert heating and cooling services in Dallas-Fort Worth. Same-day service available Monday-Friday. Call (972) 777-COOL.", "Published", "https://dfwhvac.com/contact"),
    ("/services", "Professional HVAC services in Dallas-Fort Worth. Residential & commercial air conditioning, heating, and maintenance.", "Published", "https://dfwhvac.com/services"),
    ("/faq", "Get answers to common HVAC questions \u2014 air conditioning, heating, heat pumps, preventative maintenance, and more. Expert advice from DFW HVAC, serving Dallas-Fort Worth for three generations.", "Published", "https://dfwhvac.com/faq"),
    ("/reviews", "Read 130+ 5-star reviews from real DFW HVAC customers. Family-owned HVAC contractor serving Dallas-Fort Worth since 1974. See why customers trust us.", "Published", "https://dfwhvac.com/reviews"),
    ("/request-service", "Request HVAC service from DFW HVAC. Fast response, licensed technicians, upfront pricing. Serving Dallas-Fort Worth. Call (972) 777-COOL or fill out our form.", "Published", "https://dfwhvac.com/request-service"),
    ("/estimate", "Request a free HVAC estimate from DFW HVAC. Expert heating and cooling services in Dallas-Fort Worth. Honest assessments, transparent pricing.", "Published", "https://dfwhvac.com/estimate"),
    ("/cities-served", "DFW HVAC provides professional HVAC services across the Dallas-Fort Worth metroplex. Find heating and air conditioning services in your city.", "Published", "https://dfwhvac.com/cities-served"),
    ("/privacy-policy", "Privacy Policy for DFW HVAC. Learn how we collect, use, and protect your personal information.", "Published", "https://dfwhvac.com/privacy-policy"),
    ("/terms-of-service", "Terms of Service for DFW HVAC. Review the terms and conditions governing your use of our website and HVAC services.", "Published", "https://dfwhvac.com/terms-of-service"),
    ("/recent-projects", "N/A (307 redirects to /reviews)", "Unpublished", "https://dfwhvac.com/recent-projects"),

    # Section: Service Pages
    ("SECTION", "Service Pages", "", "", ""),
    ("/services/residential/air-conditioning", "FALLBACK|Complete AC installation, repair, and replacement services for Dallas-Fort Worth homes.", "Published", "https://dfwhvac.com/services/residential/air-conditioning"),
    ("/services/residential/heating", "FALLBACK|Furnace repair, heat pump service, and heating system installation for DFW homes.", "Published", "https://dfwhvac.com/services/residential/heating"),
    ("/services/residential/preventative-maintenance", "FALLBACK|Regular maintenance to keep your HVAC system running efficiently year-round.", "Published", "https://dfwhvac.com/services/residential/preventative-maintenance"),
    ("/services/residential/indoor-air-quality", "FALLBACK|Improve your home's air quality with our specialized solutions.", "Published", "https://dfwhvac.com/services/residential/indoor-air-quality"),
    ("/services/commercial/commercial-air-conditioning", "FALLBACK|Large-scale AC systems for businesses and commercial properties in the DFW area.", "Published", "https://dfwhvac.com/services/commercial/commercial-air-conditioning"),
    ("/services/commercial/commercial-heating", "FALLBACK|Commercial heating solutions and repair to keep businesses and commercial properties comfortable.", "Published", "https://dfwhvac.com/services/commercial/commercial-heating"),
    ("/services/commercial/commercial-maintenance", "FALLBACK|Comprehensive maintenance programs that help businesses keep their HVAC systems running efficiently and avoid downtime.", "Published", "https://dfwhvac.com/services/commercial/commercial-maintenance"),

    # Section: City Pages
    ("SECTION", "City Pages (28)", "", "", ""),
    ("/cities-served/allen", "FALLBACK|Professional heating and air conditioning services in Allen, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75013.", "Published", "https://dfwhvac.com/cities-served/allen"),
    ("/cities-served/argyle", "FALLBACK|Professional heating and air conditioning services in Argyle, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76226.", "Published", "https://dfwhvac.com/cities-served/argyle"),
    ("/cities-served/arlington", "FALLBACK|Professional heating and air conditioning services in Arlington, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76001, 76002, 76012, 76013, 76014, 76015, 76016, 76017, 76018.", "Published", "https://dfwhvac.com/cities-served/arlington"),
    ("/cities-served/bedford", "FALLBACK|Professional heating and air conditioning services in Bedford, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76021, 76022.", "Published", "https://dfwhvac.com/cities-served/bedford"),
    ("/cities-served/carrollton", "FALLBACK|Professional heating and air conditioning services in Carrollton, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75006, 75007, 75010.", "Published", "https://dfwhvac.com/cities-served/carrollton"),
    ("/cities-served/colleyville", "FALLBACK|Professional heating and air conditioning services in Colleyville, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76034.", "Published", "https://dfwhvac.com/cities-served/colleyville"),
    ("/cities-served/coppell", "FALLBACK|Professional heating and air conditioning services in Coppell, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75019.", "Published", "https://dfwhvac.com/cities-served/coppell"),
    ("/cities-served/dallas", "FALLBACK|Professional heating and air conditioning services in Dallas, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75201, 75202, 75204, 75205, 75206, 75207, 75208, 75209, 75214, 75218, 75219, 75225, 75229, 75230, 75238, 75244, 75248, 75254.", "Published", "https://dfwhvac.com/cities-served/dallas"),
    ("/cities-served/denton", "FALLBACK|Professional heating and air conditioning services in Denton, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76210.", "Published", "https://dfwhvac.com/cities-served/denton"),
    ("/cities-served/euless", "FALLBACK|Professional heating and air conditioning services in Euless, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76039, 76040.", "Published", "https://dfwhvac.com/cities-served/euless"),
    ("/cities-served/farmers-branch", "FALLBACK|Professional heating and air conditioning services in Farmers Branch, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75234.", "Published", "https://dfwhvac.com/cities-served/farmers-branch"),
    ("/cities-served/flower-mound", "FALLBACK|Professional heating and air conditioning services in Flower Mound, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75022, 75028, 75077.", "Published", "https://dfwhvac.com/cities-served/flower-mound"),
    ("/cities-served/fort-worth", "FALLBACK|Professional heating and air conditioning services in Fort Worth, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76118, 76120, 76137, 76177.", "Published", "https://dfwhvac.com/cities-served/fort-worth"),
    ("/cities-served/frisco", "FALLBACK|Professional heating and air conditioning services in Frisco, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75033, 75034, 75035, 75036.", "Published", "https://dfwhvac.com/cities-served/frisco"),
    ("/cities-served/grapevine", "FALLBACK|Professional heating and air conditioning services in Grapevine, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76051.", "Published", "https://dfwhvac.com/cities-served/grapevine"),
    ("/cities-served/haslet", "FALLBACK|Professional heating and air conditioning services in Haslet, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76052.", "Published", "https://dfwhvac.com/cities-served/haslet"),
    ("/cities-served/hurst", "FALLBACK|Professional heating and air conditioning services in Hurst, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76054.", "Published", "https://dfwhvac.com/cities-served/hurst"),
    ("/cities-served/irving", "FALLBACK|Professional heating and air conditioning services in Irving, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75038, 75039, 75062, 75063.", "Published", "https://dfwhvac.com/cities-served/irving"),
    ("/cities-served/keller", "FALLBACK|Professional heating and air conditioning services in Keller, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76244, 76248.", "Published", "https://dfwhvac.com/cities-served/keller"),
    ("/cities-served/lake-dallas", "FALLBACK|Professional heating and air conditioning services in Lake Dallas, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75065.", "Published", "https://dfwhvac.com/cities-served/lake-dallas"),
    ("/cities-served/lewisville", "FALLBACK|Professional heating and air conditioning services in Lewisville, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75057, 75067.", "Published", "https://dfwhvac.com/cities-served/lewisville"),
    ("/cities-served/mansfield", "FALLBACK|Professional heating and air conditioning services in Mansfield, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76063.", "Published", "https://dfwhvac.com/cities-served/mansfield"),
    ("/cities-served/north-richland-hills", "FALLBACK|Professional heating and air conditioning services in North Richland Hills, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76148, 76180, 76182.", "Published", "https://dfwhvac.com/cities-served/north-richland-hills"),
    ("/cities-served/plano", "FALLBACK|Professional heating and air conditioning services in Plano, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75023, 75024, 75025, 75075, 75093, 75252.", "Published", "https://dfwhvac.com/cities-served/plano"),
    ("/cities-served/richardson", "FALLBACK|Professional heating and air conditioning services in Richardson, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75080, 75081.", "Published", "https://dfwhvac.com/cities-served/richardson"),
    ("/cities-served/roanoke", "FALLBACK|Professional heating and air conditioning services in Roanoke, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76262.", "Published", "https://dfwhvac.com/cities-served/roanoke"),
    ("/cities-served/southlake", "FALLBACK|Professional heating and air conditioning services in Southlake, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 76092.", "Published", "https://dfwhvac.com/cities-served/southlake"),
    ("/cities-served/the-colony", "FALLBACK|Professional heating and air conditioning services in The Colony, Texas. Same-day HVAC repair, installation, and maintenance. Serving zip codes: 75056.", "Published", "https://dfwhvac.com/cities-served/the-colony"),
]

row = 2
for page in pages:
    if page[0] == "SECTION":
        cell = ws.cell(row=row, column=1, value=page[1])
        cell.font = section_font
        cell.fill = section_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = thin_border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        continue

    url, meta, status, cms_nav = page[0], page[1], page[2], page[3]
    is_fallback = meta.startswith("FALLBACK|")
    display_meta = meta.replace("FALLBACK|", "") if is_fallback else meta

    ws.cell(row=row, column=1, value=url).alignment = wrap
    ws.cell(row=row, column=2, value=display_meta).alignment = wrap
    ws.cell(row=row, column=3, value=status).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row, column=4, value=cms_nav).alignment = wrap

    for col in range(1, 5):
        ws.cell(row=row, column=col).border = thin_border

    # Highlight issues
    if "WARNING" in meta:
        ws.cell(row=row, column=2).fill = warning_fill
    elif is_fallback:
        ws.cell(row=row, column=2).fill = null_fill

    if status == "Unpublished":
        ws.cell(row=row, column=3).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    row += 1

# Legend
row += 1
ws.cell(row=row, column=1, value="Legend:").font = Font(bold=True)
row += 1
ws.cell(row=row, column=1, value="Yellow = Needs immediate fix (test/incorrect data)")
ws.cell(row=row, column=1).fill = warning_fill
row += 1
ws.cell(row=row, column=1, value="Red = Fallback code description (no custom meta description set in CMS)")
ws.cell(row=row, column=1).fill = null_fill

wb.save("/app/frontend/public/DFW_HVAC_Site_Audit.xlsx")
print("Spreadsheet saved!")
